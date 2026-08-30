"""XLSX / XLS -> XML converter with a small Tkinter GUI.

Uruchomienie:  python xlsx2xml.py
Autotest:      python xlsx2xml.py --selftest
"""
import datetime
import json
import os
import re
import sys
import threading
import xml.etree.ElementTree as ET

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

EXTS = (".xlsx", ".xlsm", ".xls")
CFG_PATH = os.path.join(os.path.expanduser("~"), ".xlsx2xml.json")


def load_cfg():
    try:
        with open(CFG_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_cfg(data):
    try:
        with open(CFG_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _tag(name, idx):
    """Zamienia nazwe kolumny na poprawna nazwe tagu XML."""
    t = re.sub(r"\W+", "_", str(name).strip(), flags=re.UNICODE).strip("_")
    if not t:
        return "col_%d" % idx
    if not (t[0].isalpha() or t[0] == "_"):
        t = "col_" + t
    return t


def _text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


# ------------------------------------------------- odczyt arkuszy (2 formaty)

def _sheets_xlsx(path):
    if openpyxl is None:
        raise RuntimeError("Brak biblioteki openpyxl - zainstaluj: pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            yield ws.title, enumerate(ws.iter_rows(values_only=True), start=1)
    finally:
        wb.close()


def _xls_value(cell, datemode):
    # ctype: 0 empty, 1 text, 2 number, 3 date, 4 bool, 5 error, 6 blank
    if cell.ctype in (0, 5, 6):
        return None
    if cell.ctype == 3:
        return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
    if cell.ctype == 4:
        return bool(cell.value)
    v = cell.value
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def _sheets_xls(path):
    if xlrd is None:
        raise RuntimeError("Stary format .xls wymaga biblioteki xlrd - zainstaluj: pip install xlrd")
    book = xlrd.open_workbook(path)
    for sh in book.sheets():
        yield sh.name, (
            (r + 1, [_xls_value(sh.cell(r, c), book.datemode) for c in range(sh.ncols)])
            for r in range(sh.nrows)
        )


def read_sheets(path):
    """Zwraca pary (nazwa_arkusza, iterator (numer_wiersza, wartosci)) dla .xlsx/.xlsm/.xls."""
    return _sheets_xls(path) if path.lower().endswith(".xls") else _sheets_xlsx(path)


# ------------------------------------------------------------------ konwersja

def convert_file(path, out_dir, headers=True):
    """Konwertuje jeden plik Excela na XML. Zwraca sciezke pliku wynikowego."""
    root = ET.Element("workbook", name=os.path.basename(path))
    for sheet_name, rows in read_sheets(path):
        sheet_el = ET.SubElement(root, "sheet", name=sheet_name)
        names = None
        for r, row in rows:
            if all(c is None or c == "" for c in row):
                continue
            if headers and names is None:
                names = [_tag(c, i) for i, c in enumerate(row, start=1)]
                continue
            row_el = ET.SubElement(sheet_el, "row", index=str(r))
            for i, cell in enumerate(row, start=1):
                if cell is None or cell == "":
                    continue
                tag = names[i - 1] if names and i <= len(names) else "col_%d" % i
                ET.SubElement(row_el, tag).text = _text(cell)

    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, os.path.splitext(os.path.basename(path))[0] + ".xml")
    ET.indent(root, space="  ")
    ET.ElementTree(root).write(out_path, encoding="utf-8", xml_declaration=True)
    return out_path


def find_inputs(path):
    """Plik -> [plik]. Folder -> wszystkie arkusze w srodku (bez plikow tymczasowych ~$)."""
    if os.path.isfile(path):
        return [path]
    return sorted(
        os.path.join(path, f)
        for f in os.listdir(path)
        if f.lower().endswith(EXTS) and not f.startswith("~$")
    )


# ---------------------------------------------------------------------- GUI

def main():
    import tkinter as tk
    from tkinter import filedialog, ttk

    cfg = load_cfg()

    root = tk.Tk()
    root.title("XLSX -> XML")
    root.geometry("640x420")
    root.minsize(560, 380)

    src = tk.StringVar(value=cfg.get("src", ""))
    dst = tk.StringVar(value=cfg.get("dst", ""))
    headers = tk.BooleanVar(value=cfg.get("headers", True))

    def remember():
        cfg.update(src=src.get(), dst=dst.get(), headers=headers.get())
        save_cfg(cfg)

    frm = ttk.Frame(root, padding=10)
    frm.pack(fill="both", expand=True)
    frm.columnconfigure(1, weight=1)

    def _start_dir(var):
        p = var.get().strip()
        if os.path.isdir(p):
            return p
        if p and os.path.isdir(os.path.dirname(p)):
            return os.path.dirname(p)
        return cfg.get("last_dir", "")

    def _pick(var, is_dir):
        p = (filedialog.askdirectory(initialdir=_start_dir(var)) if is_dir else
             filedialog.askopenfilename(initialdir=_start_dir(var),
                                        filetypes=[("Excel", "*.xlsx *.xlsm *.xls"),
                                                   ("Wszystkie pliki", "*.*")]))
        if not p:
            return
        var.set(os.path.normpath(p))
        cfg["last_dir"] = p if is_dir else os.path.dirname(p)
        if var is src and not dst.get():
            dst.set(os.path.normpath(os.path.join(cfg["last_dir"], "xml")))
        remember()

    ttk.Label(frm, text="Plik / folder wejsciowy:").grid(row=0, column=0, sticky="w")
    ttk.Entry(frm, textvariable=src).grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 6))
    btns = ttk.Frame(frm)
    btns.grid(row=1, column=2, sticky="e", padx=(6, 0), pady=(0, 6))
    ttk.Button(btns, text="Plik...", width=9, command=lambda: _pick(src, False)).pack(side="left")
    ttk.Button(btns, text="Folder...", width=9, command=lambda: _pick(src, True)).pack(side="left", padx=(4, 0))

    ttk.Label(frm, text="Folder wyjsciowy:").grid(row=2, column=0, sticky="w")
    ttk.Entry(frm, textvariable=dst).grid(row=3, column=0, columnspan=2, sticky="ew", pady=(0, 6))
    ttk.Button(frm, text="Wybierz...", width=20, command=lambda: _pick(dst, True)
               ).grid(row=3, column=2, sticky="e", padx=(6, 0), pady=(0, 6))

    ttk.Checkbutton(frm, text="Pierwszy wiersz zawiera naglowki kolumn", variable=headers,
                    command=remember).grid(row=4, column=0, columnspan=3, sticky="w", pady=(0, 8))

    log = tk.Text(frm, height=12, wrap="word", state="disabled")
    log.grid(row=6, column=0, columnspan=3, sticky="nsew")
    frm.rowconfigure(6, weight=1)
    sb = ttk.Scrollbar(frm, command=log.yview)
    sb.grid(row=6, column=3, sticky="ns")
    log.configure(yscrollcommand=sb.set)

    def say(msg):
        log.configure(state="normal")
        log.insert("end", msg + "\n")
        log.see("end")
        log.configure(state="disabled")

    def run():
        s, d = src.get().strip(), dst.get().strip()
        if not s or not os.path.exists(s):
            say("! Wskaz istniejacy plik lub folder wejsciowy.")
            return
        if not d:
            say("! Wskaz folder wyjsciowy.")
            return
        files = find_inputs(s)
        if not files:
            say("! Nie znaleziono plikow %s." % ", ".join(EXTS))
            return
        remember()
        go.configure(state="disabled")
        say("Start: %d plik(ow) -> %s" % (len(files), d))

        def work():
            ok = 0
            for f in files:
                try:
                    out = convert_file(f, d, headers.get())
                    ok += 1
                    root.after(0, say, "  OK    %s -> %s" % (os.path.basename(f), os.path.basename(out)))
                except Exception as e:
                    root.after(0, say, "  BLAD  %s: %s" % (os.path.basename(f), e))
            root.after(0, say, "Gotowe: %d/%d." % (ok, len(files)))
            root.after(0, lambda: go.configure(state="normal"))

        threading.Thread(target=work, daemon=True).start()

    go = ttk.Button(frm, text="Konwertuj", command=run)
    go.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(0, 8))

    if openpyxl is None:
        say("! Brak biblioteki openpyxl (.xlsx). Zainstaluj:  pip install openpyxl")
    if xlrd is None:
        say("! Brak biblioteki xlrd (stary format .xls). Zainstaluj:  pip install xlrd")

    def on_close():
        remember()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()


# ------------------------------------------------------------------ selftest

def selftest():
    import tempfile
    assert _tag("Imie i nazwisko", 1) == "Imie_i_nazwisko"
    assert _tag("2024", 1) == "col_2024"
    assert _tag("   ", 3) == "col_3"
    assert _text(None) == "" and _text(True) == "true"

    tmp = tempfile.mkdtemp()
    xlsx = os.path.join(tmp, "test.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Imie", "Wiek"])
    ws.append(["Anna", 30])
    ws.append([None, None])
    ws.append(["Jan", 41])
    wb.save(xlsx)

    out = convert_file(xlsx, os.path.join(tmp, "out"))
    rows = ET.parse(out).findall(".//row")
    assert len(rows) == 2, rows
    assert rows[0].find("Imie").text == "Anna"
    assert rows[1].find("Wiek").text == "41"

    open(os.path.join(tmp, "stary.xls"), "wb").close()
    open(os.path.join(tmp, "~$tymczasowy.xlsx"), "wb").close()
    found = [os.path.basename(f) for f in find_inputs(tmp)]
    assert found == ["stary.xls", "test.xlsx"], found

    backup = load_cfg()
    save_cfg({"src": "X"})
    assert load_cfg()["src"] == "X"
    save_cfg(backup)

    print("selftest OK ->", out)


if __name__ == "__main__":
    selftest() if "--selftest" in sys.argv else main()
